"""
bot_recurso_glosa.py

Robô de extração do Recurso de Glosa Eletrônica (Portal SulAmérica, Recursar Mat/Med).
Expõe um endpoint HTTP (FastAPI) que o n8n chama uma vez por Lote.

SELETORES: os seletores de login, navegação, filtros e do modal de recursos foram
extraídos do HTML real via `playwright codegen` (não são palpites). Os pontos ainda
marcados com "CONFIRMAR" são os que dependem da estrutura interna das tabelas de
resultado — use o script inspecionar.py para capturar o HTML delas.

O sistema é JSF/PrimeFaces: os IDs contêm ":" e por isso usamos seletores de
atributo [id="..."] em vez de #id (que exigiria escapar os dois-pontos).
"""

import os
import re
import time
import logging
import threading
from datetime import date
from calendar import monthrange
from typing import Optional, List, Dict

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from playwright.sync_api import sync_playwright, Page, TimeoutError as PWTimeout

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("bot_recurso_glosa")

PORTAL_URL = os.environ.get(
    "SULAMERICA_PORTAL_URL",
    "https://saude.sulamericaseguros.com.br/prestador/login/",
)
RGE_URL = os.environ.get(
    "SULAMERICA_RGE_URL",
    "https://saude.sulamericaseguros.com.br/prestador/servicos-medicos/recurso-de-glosa-eletronico/rge/",
)
PORTAL_CODIGO = os.environ["SULAMERICA_CODIGO"]
PORTAL_USER = os.environ["SULAMERICA_USER"]
PORTAL_PASS = os.environ["SULAMERICA_PASS"]
HEADLESS = os.environ.get("HEADLESS", "true").lower() == "true"
# Esperas fixas. Sao curtas, mas se repetem centenas de vezes por execucao -
# baixar de 400 para 150 economiza minutos no total. Se o portal comecar a
# falhar por leitura antecipada, suba os valores em vez de mexer no codigo.
DELAY_MS = int(os.environ.get("DELAY_ENTRE_ACOES_MS", "150"))
ESPERA_RENDER_MS = int(os.environ.get("ESPERA_RENDER_MS", "350"))
# Imagens/fontes/rastreadores nao sao usados na extracao; bloquear acelera as
# telas. Defina BLOQUEAR_MIDIA=false para depurar vendo o portal completo.
BLOQUEAR_MIDIA = os.environ.get("BLOQUEAR_MIDIA", "true").lower() == "true"

# IDs reais confirmados via playwright codegen
ID_LOTE = '[id="formFiltros:loteConta:loteConta"]'
ID_DATA_INI = "formFiltros:dataPagamentoConta"
ID_DATA_FIM = "formFiltros:dataPagamentoContaFim"
ID_TABELA_GUIAS = "formGrid:formGrid:gridTable"
ID_CHECKBOX_DISPONIVEIS = "formFiltros:somenteGuiasDisponiveis"
ID_BTN_PESQUISAR = "formFiltros:btnPesquisar"
ID_BTN_DETALHES = "formGrid:formButtons:modalDialogButton"
MENU_POR_ABA = {"matmed": "formLink:menuMatMed", "itens": "formLink:menuItens"}
ABAS = ("matmed", "itens")

# O Mat/Med tem um seletor "Glosa Administrativa / Glosa Tecnica" que o robo
# nunca tocava - ficava sempre no padrao (administrativa). Protocolos de glosa
# tecnica existiam no portal e nunca eram encontrados. A aba Itens nao tem esse
# seletor, entao ela e percorrida uma vez so.
PASSAGENS = (
    ("matmed", "administrativa"),
    ("matmed", "tecnica"),
    ("itens", None),
)
ID_TABELA_RECURSOS = "formGrid:formRecursos:gridTableRecursos"
ID_TABELA_ITENS = "formRecursar:itensRecursoTable"
IDS_DATA_REALIZACAO = (
    "visualizarDadosGuia:dataRealizacao",     # aba Mat/Med
    "dadosGuiaContaGlosada:dataRealizacao",   # aba Itens
)

app = FastAPI(title="Bot Recurso de Glosa - SulAmérica")

# O n8n dispara varios lotes em paralelo e cada um abriria um Chromium.
# Isso esgotava a memoria da VPS e fazia o portal recusar conexoes - dai a
# enxurrada de "Page.goto: Timeout" no login. Uma execucao por vez resolve.
_EXECUCOES_SIMULTANEAS = int(os.environ.get("MAX_EXECUCOES_SIMULTANEAS", "1"))
_semaforo = threading.Semaphore(_EXECUCOES_SIMULTANEAS)


class SemResultados(Exception):
    """O portal respondeu "Nenhum item disponivel" - a busca funcionou, o lote
    e que nao tem nada nesta aba. Diferente de a busca ter quebrado."""



class ProcessarLoteRequest(BaseModel):
    lote: str
    data_inicio_pagto: str   # dd/mm/aaaa
    data_fim_pagto: str      # dd/mm/aaaa


class ItemExtraido(BaseModel):
    lote: str = ""
    aba: str = ""
    guia: str = ""
    protocolo: str = ""
    data_recurso: Optional[str] = None
    data_complemento: Optional[str] = None
    valor_unit: Optional[str] = None
    valor_total: Optional[str] = None
    valor_acatado: Optional[str] = None
    qtd_itens_protocolo: Optional[str] = None
    data_uso: Optional[str] = None
    descricao_item: Optional[str] = None
    codigo_item: Optional[str] = None
    cod_glosa: Optional[str] = None
    qtde: Optional[int] = None
    justificativa: Optional[str] = None
    revisao_manual: bool = False
    erro: Optional[str] = None


# ------------------------------------------------------------- utilitários ----

def calcular_range_mes(data_iso: str) -> tuple:
    """'2025-03-17' -> ('01/03/2025', '31/03/2025')"""
    d = date.fromisoformat(data_iso)
    ultimo = monthrange(d.year, d.month)[1]
    return (
        date(d.year, d.month, 1).strftime("%d/%m/%Y"),
        date(d.year, d.month, ultimo).strftime("%d/%m/%Y"),
    )


def pausa_humana(page: Page):
    page.wait_for_timeout(DELAY_MS)


def aguardar_pagina_pronta(page: Page, timeout: int = 20000):
    """Evita 'networkidle', que nunca acontece em sites com analytics rodando."""
    try:
        page.wait_for_load_state("domcontentloaded", timeout=timeout)
    except PWTimeout:
        pass
    page.wait_for_timeout(ESPERA_RENDER_MS)


def aguardar_ajax(page: Page, timeout: int = 25000):
    """O sistema usa jQuery blockUI: enquanto processa, cobre a tela com um
    overlay. Esperar ele sumir evita ler a pagina no meio do carregamento."""
    try:
        page.wait_for_function(
            "() => document.querySelectorAll('.blockUI.blockOverlay').length === 0",
            timeout=timeout,
        )
    except Exception:
        pass
    page.wait_for_timeout(ESPERA_RENDER_MS)


def capturar_screenshot_erro(page: Page, nome: str) -> str:
    import time
    caminho = f"debug_{nome}_{int(time.time())}.png"
    try:
        page.screenshot(path=caminho, full_page=True)
    except Exception:
        pass
    return caminho


def fechar_banner_cookies(page: Page):
    try:
        page.get_by_role("button", name=re.compile("^Continuar$", re.I)).click(timeout=4000)
    except Exception:
        pass


def _localizar_input_data(page: Page, base_id: str):
    """Devolve (locator, seletor) do input de texto do calendário PrimeFaces."""
    for seletor in [f'[id="{base_id}_input"]', f'[id="{base_id}"] input']:
        try:
            loc = page.locator(seletor).first
            if loc.count() > 0:
                return loc, seletor
        except Exception:
            continue
    return None, None


def preencher_data_primefaces(page: Page, base_id: str, valor: str):
    """Os campos de data são calendários PrimeFaces. Tentamos escrever direto no
    input (bem mais estável que navegar pelo widget) e SEMPRE conferimos se o
    valor realmente ficou lá — falha silenciosa aqui faz a busca voltar vazia
    sem ninguém perceber."""
    campo, seletor = _localizar_input_data(page, base_id)
    if campo is None:
        raise RuntimeError(f"Não encontrei o campo de data '{base_id}'")

    def conferir() -> bool:
        try:
            return campo.input_value(timeout=2000).strip() == valor
        except Exception:
            return False

    # 1) fill direto
    try:
        campo.fill(valor, timeout=4000)
        page.keyboard.press("Escape")
        if conferir():
            return
    except Exception:
        pass

    # 2) clicar, limpar e digitar como um humano
    try:
        campo.click(timeout=4000)
        # ControlOrMeta: o robo roda em Linux (Ctrl), nao em macOS (Meta).
        # Com "Meta+A" fixo esse fallback nunca limpava o campo em producao.
        page.keyboard.press("ControlOrMeta+A")
        page.keyboard.press("Delete")
        page.keyboard.type(valor, delay=60)
        page.keyboard.press("Escape")
        if conferir():
            return
    except Exception:
        pass

    # 3) via JavaScript (funciona mesmo se o input for readonly)
    try:
        page.evaluate(
            """([sel, val]) => {
                const el = document.querySelector(sel);
                if (!el) return false;
                el.removeAttribute('readonly');
                el.value = val;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }""",
            [seletor, valor],
        )
        if conferir():
            return
    except Exception:
        pass

    atual = ""
    try:
        atual = campo.input_value()
    except Exception:
        pass
    raise RuntimeError(
        f"Não consegui preencher a data em '{base_id}'. "
        f"Esperado '{valor}', campo ficou com '{atual}'"
    )


# -------------------------------------------------------- passos do fluxo ----

def fazer_login(page: Page, tentativas: int = 3):
    """O portal fica intermitente sob carga. Em vez de falhar de primeira,
    tenta de novo com espera crescente."""
    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            page.goto(PORTAL_URL, timeout=60000, wait_until="domcontentloaded")
            ultimo_erro = None
            break
        except Exception as e:
            ultimo_erro = e
            espera = 3000 * tentativa
            log.warning("Login tentativa %d/%d falhou (%s); aguardando %dms",
                        tentativa, tentativas, type(e).__name__, espera)
            page.wait_for_timeout(espera)
    if ultimo_erro:
        raise RuntimeError(f"Portal inacessivel apos {tentativas} tentativas: {ultimo_erro}")

    aguardar_pagina_pronta(page)
    fechar_banner_cookies(page)
    try:
        page.locator("#code").fill(PORTAL_CODIGO, timeout=8000)
        page.locator("#user").fill(PORTAL_USER, timeout=8000)
        page.locator("#senha").fill(PORTAL_PASS, timeout=8000)
        page.get_by_role("button", name=re.compile("^Entrar$", re.I)).click()
        aguardar_pagina_pronta(page)
    except Exception as e:
        caminho = capturar_screenshot_erro(page, "login")
        raise RuntimeError(f"Falha no login ({e}). Print salvo em: {caminho}")


def abrir_rge(page: Page) -> Page:
    """O sistema RGE abre numa janela separada (popup). Devolve essa nova página."""
    try:
        with page.expect_popup(timeout=20000) as info:
            page.goto(RGE_URL)
        rge = info.value
    except Exception:
        # Em alguns casos pode abrir na mesma aba em vez de popup
        rge = page
    aguardar_pagina_pronta(rge)
    return rge


def ir_para_aba(rge: Page, aba: str):
    """As duas abas (Mat/Med e Itens) usam exatamente os mesmos IDs de campo e
    de tabela - confirmado inspecionando o HTML das duas. So muda o caminho de
    entrada e o estado inicial do checkbox."""
    # Ao terminar uma aba, o modal "Detalhes da Guia" costuma ficar aberto e
    # o overlay dele bloqueia o clique no menu da outra aba.
    fechar_modal(rge)

    menu = MENU_POR_ABA[aba]
    rotulo = "Recursar Mat/Med" if aba == "matmed" else "Recursar Itens"
    tentativas = [
        ("menu do topo", lambda: rge.locator(f'[id="{menu}"]').click(timeout=10000)),
        ("link da tela inicial", lambda: rge.get_by_role("link", name=rotulo).click(timeout=8000)),
        ("menu via JavaScript", lambda: rge.evaluate(
            "(id) => { const el = document.getElementById(id); if (!el) throw new Error('menu nao encontrado'); el.click(); }",
            menu)),
    ]

    erros = []
    for descricao, acao in tentativas:
        try:
            acao()
            log.info("Entrei na aba %s pelo %s", aba, descricao)
            break
        except Exception as e:
            erros.append(f"{descricao}: {type(e).__name__}")
            continue
    else:
        caminho = capturar_screenshot_erro(rge, f"aba_{aba}")
        raise RuntimeError(
            f"Nao consegui abrir a aba {aba}. Tentativas: {'; '.join(erros)}. Print: {caminho}"
        )

    try:  # modal de aviso com botao OK
        rge.get_by_role("button", name=re.compile("^OK$", re.I)).click(timeout=6000)
    except Exception:
        pass
    aguardar_ajax(rge)
    aguardar_pagina_pronta(rge)
    fechar_modal(rge)   # comunicado do Mat/Med bloqueia os cliques seguintes


# Mantido para o inspecionar.py continuar funcionando
def ir_para_matmed(rge: Page):
    ir_para_aba(rge, "matmed")


def fechar_modal(rge: Page):
    """Fecha QUALQUER dialogo do PrimeFaces que esteja aberto.

    Ao entrar no Mat/Med o portal exibe um modal de comunicado
    (modalComunicadoRecursarMatMed) cujo overlay intercepta os cliques. Como
    ele nao usa o mesmo botao dos demais, o robo ficava 30s tentando clicar em
    "Pesquisar" e desistia - e, por ser tratado como "aba sem resultados", a
    falha passava despercebida."""
    for _ in range(3):
        try:
            aberto = rge.evaluate(
                """() => Array.from(document.querySelectorAll('.ui-widget-overlay'))
                        .some(el => el.offsetParent !== null ||
                                    getComputedStyle(el).display !== 'none')"""
            )
        except Exception:
            aberto = False
        if not aberto:
            return

        # a) botoes de fechar / confirmar visiveis
        for seletor in (
            ".ui-dialog:visible .ui-dialog-titlebar-close",
            ".ui-dialog:visible button:has-text('OK')",
            ".ui-dialog:visible button:has-text('Fechar')",
            ".ui-dialog:visible button:has-text('Continuar')",
            ".ui-dialog:visible .ui-button",
        ):
            try:
                alvo = rge.locator(seletor)
                if alvo.count():
                    alvo.first.click(timeout=2500)
                    rge.wait_for_timeout(400)
                    break
            except Exception:
                continue
        else:
            # b) ultimo recurso: esconder dialogo e overlay via JavaScript
            try:
                rge.evaluate(
                    """() => {
                        document.querySelectorAll('.ui-widget-overlay').forEach(el => el.remove());
                        document.querySelectorAll('.ui-dialog').forEach(el => {
                            if (getComputedStyle(el).display !== 'none') el.style.display = 'none';
                        });
                    }"""
                )
                rge.wait_for_timeout(300)
            except Exception:
                pass

        try:
            rge.keyboard.press("Escape")
            rge.wait_for_timeout(200)
        except Exception:
            pass


def garantir_tela_de_pesquisa(rge: Page, aba: str = "matmed"):
    """Abrir um protocolo tira o robo da tela de pesquisa (e um submit JSF).
    Antes de cada nova busca, garante que estamos de volta nela."""
    fechar_modal(rge)
    if rge.locator(ID_LOTE).count() > 0:
        return

    for seletor in [f'[id="{MENU_POR_ABA[aba]}"]', "#menu2"]:
        try:
            rge.locator(seletor).first.click(timeout=8000)
            aguardar_ajax(rge)
            aguardar_pagina_pronta(rge)
            try:  # pode reaparecer o aviso com botao OK
                rge.get_by_role("button", name=re.compile("^OK$", re.I)).click(timeout=4000)
                aguardar_ajax(rge)
            except Exception:
                pass
            if rge.locator(ID_LOTE).count() > 0:
                return
        except Exception:
            continue

    # Ultimo recurso: reentrar pela aba do zero
    try:
        ir_para_aba(rge, aba)
        if rge.locator(ID_LOTE).count() > 0:
            return
    except Exception:
        pass

    caminho = capturar_screenshot_erro(rge, "voltar_pesquisa")
    raise RuntimeError(
        f"Nao consegui voltar para a tela de pesquisa da aba {aba}. Print: {caminho}"
    )


def preencher_lote(rge: Page, lote: str):
    """Limpa e preenche o campo Lote, CONFERINDO o valor depois. Sem limpar
    antes, o valor novo pode grudar em restos do anterior (ex: pedir
    6300000641 e o portal receber 630000064163), e a busca volta vazia sem
    deixar claro o motivo."""
    campo = rge.locator(ID_LOTE)
    campo.wait_for(timeout=15000)

    def conferir() -> bool:
        try:
            return campo.input_value(timeout=3000).strip() == lote
        except Exception:
            return False

    # 1) limpar e preencher
    try:
        campo.fill("", timeout=5000)
        rge.wait_for_timeout(150)
        campo.fill(lote, timeout=5000)
        if conferir():
            return
    except Exception:
        pass

    # 2) selecionar tudo e digitar por cima
    try:
        campo.click(timeout=5000)
        rge.keyboard.press("ControlOrMeta+A")
        rge.keyboard.press("Delete")
        rge.keyboard.type(lote, delay=40)
        if conferir():
            return
    except Exception:
        pass

    # 3) via JavaScript, disparando os eventos que o JSF escuta
    try:
        rge.evaluate(
            """([sel, val]) => {
                const el = document.querySelector(sel);
                if (!el) return;
                el.value = '';
                el.value = val;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            }""",
            [ID_LOTE, lote],
        )
        if conferir():
            return
    except Exception:
        pass

    atual = ""
    try:
        atual = campo.input_value()
    except Exception:
        pass
    raise RuntimeError(f"Campo Lote ficou com '{atual}' em vez de '{lote}'")


def selecionar_tipo_glosa(rge: Page, tipo: Optional[str]) -> None:
    """Marca "Glosa Administrativa" ou "Glosa Tecnica" no topo da tela.

    Precisa ser feito ANTES de preencher os filtros: trocar o radio dispara um
    recarregamento do PrimeFaces que limpa lote e datas."""
    if not tipo:
        return

    rotulo = "Glosa Técnica" if tipo == "tecnica" else "Glosa Administrativa"

    def marcado() -> str:
        try:
            return rge.evaluate(
                """() => {
                    for (const inp of document.querySelectorAll('input[type=radio]')) {
                        const marcado = inp.checked ||
                            (inp.closest('.ui-radiobutton') || {}).querySelector?.('.ui-state-active');
                        if (!marcado) continue;
                        const caixa = inp.closest('td, div, span, label');
                        const texto = (caixa && caixa.parentElement)
                            ? caixa.parentElement.innerText : '';
                        if (/t[ée]cnica/i.test(texto)) return 'tecnica';
                        if (/administrativa/i.test(texto)) return 'administrativa';
                    }
                    return '';
                }"""
            )
        except Exception:
            return ""

    for seletor in (
        f'label:has-text("{rotulo}")',
        f'td:has-text("{rotulo}") .ui-radiobutton-box',
        f'span:has-text("{rotulo}")',
    ):
        try:
            alvo = rge.locator(seletor)
            if not alvo.count():
                continue
            alvo.first.click(timeout=5000)
            aguardar_ajax(rge)
            aguardar_pagina_pronta(rge)
            # VERIFICA se o radio realmente mudou. O clique pode "funcionar"
            # (sem excecao) e o PrimeFaces ignorar - e ai a passagem tecnica
            # roda como administrativa: todos os protocolos ja estao em
            # "vistos" e o resultado e zero novos, indistinguivel de "lote
            # sem glosa tecnica". Foi a busca tecnica que recuperou 70% dos
            # protocolos perdidos; falha silenciosa aqui nao pode existir.
            atual = marcado()
            if atual == tipo:
                log.info("Tipo de glosa: %s (confirmado)", rotulo)
                return
            if atual:
                log.warning("Cliquei em '%s' mas o radio ficou em '%s'; "
                            "tentando outra estrategia", rotulo, atual)
                continue
            # estado ilegivel (layout mudou?): aceita o clique com aviso,
            # comportamento antigo, para nao quebrar por falha de leitura
            log.warning("Tipo de glosa: %s (clique feito, estado nao legivel)", rotulo)
            return
        except Exception:
            continue

    # Ultimo recurso: clicar no radio pelo texto vizinho, via JavaScript
    try:
        ok = rge.evaluate(
            """(rotulo) => {
                for (const inp of document.querySelectorAll('input[type=radio]')) {
                    const bloco = inp.closest('td, div, span, label');
                    const texto = (bloco && bloco.parentElement)
                        ? bloco.parentElement.innerText : '';
                    if (texto && texto.includes(rotulo)) {
                        const caixa = (inp.closest('.ui-radiobutton') || {})
                            .querySelector?.('.ui-radiobutton-box');
                        (caixa || inp).click();
                        return true;
                    }
                }
                return false;
            }""",
            rotulo,
        )
        if ok:
            aguardar_ajax(rge)
            aguardar_pagina_pronta(rge)
            atual = marcado()
            if atual == tipo:
                log.info("Tipo de glosa: %s (via JavaScript, confirmado)", rotulo)
                return
            if not atual:
                log.warning("Tipo de glosa: %s (via JavaScript, estado nao legivel)", rotulo)
                return
    except Exception:
        pass

    # Nenhuma estrategia confirmou o tipo pedido: abortar a passagem em vez
    # de pesquisar com o filtro errado. O chamador (pesquisar_lote ->
    # loop de passagens) converte isso em linha REVISAR na planilha.
    caminho = capturar_screenshot_erro(rge, f"tipo_glosa_{tipo}")
    raise RuntimeError(
        f"Nao consegui selecionar '{rotulo}' (radio nao mudou). Print: {caminho}"
    )


def desmarcar_somente_disponiveis(rge: Page):
    """Com esse filtro ligado, guias ja recursadas somem do resultado - e e
    justamente o historico que queremos. No Mat/Med vem desmarcado; na aba
    Itens vem MARCADO por padrao.

    A falha aqui NAO pode ser silenciosa: pesquisar com o filtro ligado
    devolve so as guias ainda disponiveis para recurso, e todo o historico
    (recursadas, expiradas) some sem nenhum vestigio na planilha."""

    def estado():
        """True/False = estado lido; None = nao consegui ler."""
        try:
            return rge.locator(f'[id="{ID_CHECKBOX_DISPONIVEIS}"]').is_checked(timeout=3000)
        except Exception:
            pass
        try:
            return rge.evaluate(
                "(id) => { const el = document.getElementById(id); "
                "return el ? el.checked : null; }",
                ID_CHECKBOX_DISPONIVEIS,
            )
        except Exception:
            return None

    atual = estado()
    if atual is None:
        # Campo nao encontrado/ilegivel: pode ser mudanca de layout. Segue
        # com aviso (comportamento antigo) - a checagem de resultado vazio
        # em pesquisar_lote ainda protege contra a maior parte do estrago.
        log.warning("Nao consegui ler o checkbox 'somente guias disponiveis'; seguindo")
        return
    if atual is False:
        return  # ja esta como queremos

    # 1) uncheck normal do Playwright
    try:
        rge.locator(f'[id="{ID_CHECKBOX_DISPONIVEIS}"]').uncheck(timeout=6000)
        pausa_humana(rge)
    except Exception as e:
        log.warning("uncheck() falhou (%s); tentando via JavaScript", e)
    if estado() is False:
        return

    # 2) fallback: clicar via JavaScript (cobre o caso do input escondido
    # atras do widget do PrimeFaces) e disparar o evento que o JSF escuta
    try:
        rge.evaluate(
            """(id) => {
                const el = document.getElementById(id);
                if (!el) return;
                const caixa = (el.closest('.ui-chkbox') || {})
                    .querySelector?.('.ui-chkbox-box');
                if (caixa) { caixa.click(); return; }
                el.click();
            }""",
            ID_CHECKBOX_DISPONIVEIS,
        )
        aguardar_ajax(rge)
    except Exception:
        pass
    if estado() is False:
        return

    caminho = capturar_screenshot_erro(rge, "checkbox_disponiveis")
    raise RuntimeError(
        "Checkbox 'somente guias disponiveis' continua MARCADO apos 2 "
        f"tentativas - pesquisar assim esconderia o historico. Print: {caminho}"
    )


def pesquisar_lote(rge: Page, lote: str, data_inicio: str, data_fim: str,
                   aba: str = "matmed", tipo_glosa: Optional[str] = None):
    try:
        garantir_tela_de_pesquisa(rge, aba)
        # Antes dos filtros: trocar o radio recarrega o formulario
        selecionar_tipo_glosa(rge, tipo_glosa)
        preencher_lote(rge, lote)

        # CRÍTICO: com esse checkbox marcado, guias já recursadas NÃO aparecem.
        # Como buscamos o histórico de recursos já enviados, ele precisa ficar
        # desmarcado. Fazemos isso ANTES das datas porque esse clique dispara
        # um recarregamento parcial do PrimeFaces que pode resetar os campos.
        desmarcar_somente_disponiveis(rge)

        preencher_data_primefaces(rge, ID_DATA_INI, data_inicio)
        preencher_data_primefaces(rge, ID_DATA_FIM, data_fim)
        log.info("Filtros preenchidos: lote=%s, %s a %s", lote, data_inicio, data_fim)

        pausa_humana(rge)
        executar_pesquisa(rge)

        if contar_linhas_guias(rge) == 0:
            estado = ler_estado_filtros(rge)
            filtros_ok = str(estado.get("lote", "")).strip() == str(lote).strip()
            portal_disse_vazio = "nenhum item" in str(estado.get("mensagem_tabela", "")).lower()

            # Se os filtros estao certos e o portal ja respondeu "nenhum item",
            # o lote realmente nao tem nada nesta aba. Refazer a busca so para
            # confirmar custava ~30s por lote - e lotes sem Mat/Med sao comuns.
            if filtros_ok and portal_disse_vazio:
                raise SemResultados(
                    f"Lote {lote} nao tem guias na aba {aba} (filtros conferidos)"
                )

            # Sem a mensagem de vazio, pode ter sido ajax perdido: vale repetir.
            log.warning("Busca do lote %s sem resposta clara; tentando novamente", lote)
            rge.wait_for_timeout(1000)
            executar_pesquisa(rge)

        if contar_linhas_guias(rge) == 0:
            estado = ler_estado_filtros(rge)
            if (str(estado.get("lote", "")).strip() == str(lote).strip()
                    and "nenhum item" in str(estado.get("mensagem_tabela", "")).lower()):
                raise SemResultados(
                    f"Lote {lote} nao tem guias na aba {aba} (filtros conferidos)"
                )
            caminho = capturar_screenshot_erro(rge, "busca_vazia")
            raise RuntimeError(
                f"Busca do lote {lote} na aba {aba} nao retornou linhas. "
                f"Filtros como o portal recebeu: {estado}. Print: {caminho}"
            )
    except SemResultados:
        # "lote sem guias nesta aba" e situacao normal: precisa passar direto,
        # senao vira RuntimeError e acaba gravado como linha de REVISAR no
        # Google Sheets com a mensagem enganosa "Busca falhou".
        raise
    except Exception as e:
        caminho = capturar_screenshot_erro(rge, "pesquisa")
        raise RuntimeError(f"Falha na pesquisa do lote {lote} ({e}). Print: {caminho}")


def ler_estado_filtros(rge: Page) -> Dict:
    """Le de volta o que REALMENTE ficou nos filtros. Serve para diagnosticar
    buscas que voltam vazias: mostra se o valor digitado foi aceito pelo
    PrimeFaces ou se algo reverteu para o padrao."""
    estado = {}
    try:
        estado["lote"] = rge.locator(ID_LOTE).input_value(timeout=3000)
    except Exception:
        estado["lote"] = "<nao lido>"
    for rotulo, base in (("data_inicio", ID_DATA_INI), ("data_fim", ID_DATA_FIM)):
        try:
            campo, _ = _localizar_input_data(rge, base)
            estado[rotulo] = campo.input_value(timeout=3000) if campo else "<nao encontrado>"
        except Exception:
            estado[rotulo] = "<nao lido>"
    try:
        chk = rge.get_by_role("checkbox", name=re.compile("Exibir somente as guias", re.I))
        estado["checkbox_somente_disponiveis"] = chk.is_checked(timeout=3000)
    except Exception:
        estado["checkbox_somente_disponiveis"] = "<nao lido>"
    try:
        vazio = rge.locator(f'[id="{ID_TABELA_GUIAS}_data"] tr.ui-datatable-empty-message')
        estado["mensagem_tabela"] = vazio.first.inner_text(timeout=2000) if vazio.count() else ""
    except Exception:
        estado["mensagem_tabela"] = "<nao lido>"
    return estado


def contar_linhas_guias(rge: Page) -> int:
    try:
        return rge.locator(f'[id="{ID_TABELA_GUIAS}_data"] tr[data-ri]').count()
    except Exception:
        return 0


def executar_pesquisa(rge: Page, timeout: int = 40000) -> bool:
    """Clica em Pesquisar e espera o conteudo da tabela REALMENTE mudar.

    Antes, a espera terminava assim que encontrava qualquer linha - inclusive a
    mensagem "nenhum item" que sobrou da busca anterior. O robo lia zero
    resultados e refazia a busca: toda pesquisa custava o dobro."""
    try:
        antes = rge.evaluate(
            """(id) => {
                const t = document.getElementById(id + '_data');
                return t ? t.innerHTML.length : -1;
            }""",
            ID_TABELA_GUIAS,
        )
    except Exception:
        antes = -1

    rge.locator(f'[id="{ID_BTN_PESQUISAR}"]').click()

    try:
        rge.wait_for_function(
            """([id, antes]) => {
                const t = document.getElementById(id + '_data');
                if (!t) return false;
                const carregando = document.querySelectorAll('.blockUI.blockOverlay').length > 0;
                return !carregando && t.innerHTML.length !== antes;
            }""",
            arg=[ID_TABELA_GUIAS, antes],
            timeout=timeout,
        )
        mudou = True
    except Exception:
        # Pode ser que o resultado seja identico ao anterior - nao e erro
        mudou = False

    aguardar_ajax(rge)
    aguardar_pagina_pronta(rge)
    return mudou


def esperar_tabela_guias(rge: Page, timeout: int = 30000):
    """Espera a tabela de resultados terminar de renderizar - seja com linhas,
    seja com a mensagem de 'nenhum item'. Sem isso, ler a tabela cedo demais
    faz parecer que a busca nao retornou nada."""
    seletor = (
        f'[id="{ID_TABELA_GUIAS}_data"] tr[data-ri], '
        f'[id="{ID_TABELA_GUIAS}_data"] tr.ui-datatable-empty-message'
    )
    try:
        rge.locator(seletor).first.wait_for(timeout=timeout)
    except Exception:
        log.warning("Tabela de guias nao renderizou dentro do tempo esperado")


def paginas_de_resultado(rge: Page) -> int:
    """Quantas paginas o resultado da pesquisa tem (1 = sem paginacao).
    Le apenas o primeiro paginador: a tabela renderiza um no topo e outro no
    rodape com os mesmos botoes, e contar os dois dobraria o numero."""
    try:
        pag = rge.locator(".ui-paginator").first
        if not pag.count():
            return 1
        n = pag.locator(".ui-paginator-page").count()
        return max(1, n)
    except Exception:
        return 1


def listar_guias(rge: Page) -> List[Dict]:
    """Tabela de resultados. As celulas tem classes nomeadas (grid-lote,
    grid-guia, grid-paciente...), confirmadas no HTML real - bem mais
    estaveis do que indices de coluna."""
    # O padrao do portal e 10 guias por pagina, e ler so a pagina atual
    # perderia as demais EM SILENCIO. Quando ha mais de uma pagina, sobe o
    # paginador para 50 (o maior valor do dropdown) antes de ler. Se mesmo
    # assim sobrar paginacao (>50 guias), o chamador detecta via
    # paginas_de_resultado() e grava uma linha REVISAR na planilha.
    try:
        if paginas_de_resultado(rge) > 1:
            resultado = rge.evaluate(
                """() => {
                    const sel = document.querySelector('select.ui-paginator-rpp-options');
                    if (!sel) return 'sem_select';
                    const opcoes = Array.from(sel.options).map(o => o.value);
                    const alvo = opcoes.includes('50') ? '50' : opcoes[opcoes.length - 1];
                    if (sel.value === alvo) return 'ja_estava';
                    sel.value = alvo;
                    sel.dispatchEvent(new Event('change', { bubbles: true }));
                    return 'ok:' + alvo;
                }"""
            )
            if str(resultado).startswith("ok"):
                aguardar_ajax(rge)
                esperar_tabela_guias(rge)
                log.info("Resultado paginado: paginador ajustado (%s)", resultado)
            else:
                log.warning("Resultado paginado; ajuste do paginador: %s", resultado)
    except Exception as e:
        log.warning("Nao consegui ajustar o paginador para 50: %s", e)

    if paginas_de_resultado(rge) > 1:
        log.warning("Resultado ainda tem %d paginas; apenas a atual sera lida",
                    paginas_de_resultado(rge))

    linhas = rge.locator(f'[id="{ID_TABELA_GUIAS}_data"] tr[data-ri]')
    guias = []
    for i in range(linhas.count()):
        linha = linhas.nth(i)
        if "ui-datatable-empty-message" in (linha.get_attribute("class") or ""):
            continue

        def celula(classe: str) -> str:
            try:
                return linha.locator(f"td.{classe}").first.inner_text().strip()
            except Exception:
                return ""

        guias.append({
            "row_index": i,
            "lote": celula("grid-lote"),
            "guia": celula("grid-guia"),
            "paciente": celula("grid-paciente"),
            "data_pagamento": celula("grid-pgto"),
            "data_atendimento": celula("grid-atend"),
        })
    return guias


def indice_da_guia(rge: Page, numero_guia: str) -> Optional[int]:
    """Acha a posicao da linha cuja coluna 'Guia' bate com o numero dado.

    O row_index e capturado na pesquisa original; se a busca for refeita e o
    portal devolver as linhas em outra ordem, o indice antigo aponta para
    outra guia - e os protocolos dela seriam gravados sob o rotulo errado.
    Devolve None quando nao encontra (ou nao consegue ler)."""
    alvo = str(numero_guia or "").strip()
    if not alvo:
        return None
    try:
        linhas = rge.locator(f'[id="{ID_TABELA_GUIAS}_data"] tr[data-ri]')
        for i in range(linhas.count()):
            try:
                atual = linhas.nth(i).locator("td.grid-guia").first.inner_text().strip()
            except Exception:
                continue
            if atual == alvo:
                return i
    except Exception:
        pass
    return None


def abrir_detalhes_guia(rge: Page, row_index: int, numero_guia: str = ""):
    # O modal da guia anterior fica aberto e seu overlay intercepta o clique
    # na proxima linha. Era a causa das falhas em cascata: a primeira guia do
    # lote funcionava e todas as seguintes davam timeout de 30s.
    fechar_modal(rge)
    esperar_tabela_guias(rge)

    # Se soubermos o numero da guia, confirmamos a posicao dela na tabela
    # atual em vez de confiar no indice da busca anterior.
    if numero_guia:
        atual = indice_da_guia(rge, numero_guia)
        if atual is None:
            raise RuntimeError(
                f"Guia {numero_guia} nao esta na tabela de resultados atual"
            )
        if atual != row_index:
            log.warning("Guia %s mudou de posicao (%d -> %d) apos nova busca",
                        numero_guia, row_index, atual)
        row_index = atual

    linhas = rge.locator(f'[id="{ID_TABELA_GUIAS}_data"] tr[data-ri]')
    linha = linhas.nth(row_index)
    linha.wait_for(timeout=30000)
    linha.click()
    pausa_humana(rge)
    # Captura o conteudo ANTES do clique para poder confirmar depois que a
    # tabela realmente trocou - nao so que o elemento existe. O PrimeFaces
    # reaproveita o mesmo modal entre aberturas (id igual), entao
    # `.wait_for()` num elemento que ja existia antes retorna na hora, sem
    # esperar o AJAX preencher o conteudo novo. Isso deixava a leitura
    # vulneravel a pegar a tabela da guia/tipo ANTERIOR - mesma familia do
    # bug que o executar_pesquisa ja resolve para a lista de guias (comparar
    # innerHTML antes/depois), nunca replicada aqui. E o motivo mais provavel
    # de protocolos que so existem sob um tipo de glosa especifico (ex.:
    # Tecnica) sumirem quando a guia ja tinha sido aberta antes sob outro tipo.
    try:
        antes = rge.evaluate(
            """(id) => {
                const t = document.getElementById(id + '_data');
                return t ? t.innerHTML.length : -1;
            }""",
            ID_TABELA_RECURSOS,
        )
    except Exception:
        antes = -1

    rge.locator(f'[id="{ID_BTN_DETALHES}"]').click()
    aguardar_ajax(rge)
    rge.locator(f'[id="{ID_TABELA_RECURSOS}_data"]').wait_for(timeout=20000)

    try:
        rge.wait_for_function(
            """([id, antes]) => {
                const t = document.getElementById(id + '_data');
                return t && t.innerHTML.length !== antes;
            }""",
            arg=[ID_TABELA_RECURSOS, antes],
            timeout=8000,
        )
    except Exception:
        # Pode ser coincidencia genuina (mesma quantidade de caracteres) ou
        # atraso real do portal - registra para investigar sem travar o
        # lote inteiro por causa disso.
        log.warning(
            "Tabela de protocolos nao mudou de conteudo apos abrir a guia "
            "(guia=%s); pode estar lendo dado da abertura anterior",
            numero_guia or row_index,
        )

    pausa_humana(rge)


def _container_datatable(rge: Page, id_tabela: str):
    return rge.locator(f'[id="{id_tabela}_data"]').locator(
        'xpath=ancestor::div[contains(@class,"ui-datatable")][1]'
    )


def paginas_do_modal(rge: Page, id_tabela: str) -> int:
    """Quantas paginas a tabela DENTRO do modal atual tem (1 = sem paginacao
    ou nao encontrada). Usa o mesmo container do id_tabela, nao o paginador
    da lista de guias por tras do modal."""
    try:
        container = _container_datatable(rge, id_tabela)
        if not container.count():
            return 1
        paginador = container.locator(".ui-paginator").first
        if not paginador.count():
            return 1
        return max(1, paginador.locator(".ui-paginator-page").count())
    except Exception:
        return 1


def expandir_paginador_do_modal(rge: Page, id_tabela: str) -> None:
    """Sobe para 50 por pagina o paginador de DENTRO do modal 'Detalhes da
    Guia' (a lista de protocolos de uma guia), se ele existir e tiver mais de
    uma pagina.

    Este e um paginador DIFERENTE do da lista de guias da pesquisa (esse ja e
    tratado em listar_guias/paginas_de_resultado).

    IMPORTANTE: nao usa select_option() do Playwright. O PrimeFaces costuma
    esconder o <select> nativo atras de um widget estilizado, e o Playwright
    reprova acoes em elemento invisivel - a excecao caia no except, virava so
    um warning no log, e o robo seguia lendo apenas a pagina 1. Setar o valor
    e disparar o evento 'change' via JavaScript funciona com o select
    escondido ou visivel."""
    try:
        n_paginas = paginas_do_modal(rge, id_tabela)
        if n_paginas <= 1:
            return

        resultado = rge.evaluate(
            """(idTabela) => {
                const tbody = document.getElementById(idTabela + '_data');
                if (!tbody) return 'sem_tabela';
                const dt = tbody.closest('div.ui-datatable');
                if (!dt) return 'sem_datatable';
                const sel = dt.querySelector('select.ui-paginator-rpp-options');
                if (!sel) return 'sem_select';
                const opcoes = Array.from(sel.options).map(o => o.value);
                // pega a maior opcao disponivel (normalmente 50)
                const alvo = opcoes.includes('50') ? '50'
                           : opcoes[opcoes.length - 1];
                if (sel.value === alvo) return 'ja_estava';
                sel.value = alvo;
                sel.dispatchEvent(new Event('change', { bubbles: true }));
                return 'ok:' + alvo;
            }""",
            id_tabela,
        )
        if str(resultado).startswith("ok"):
            aguardar_ajax(rge)
            rge.locator(f'[id="{id_tabela}_data"]').wait_for(timeout=10000)
            n_depois = paginas_do_modal(rge, id_tabela)
            log.info("Modal de protocolos: paginador ajustado (%s); paginas: %d -> %d",
                     resultado, n_paginas, n_depois)
        else:
            log.warning("Modal de protocolos com %d paginas; ajuste do paginador: %s",
                        n_paginas, resultado)
    except Exception as e:
        log.warning("Nao consegui ajustar o paginador do modal de protocolos: %s", e)


def garantir_pagina_do_modal(rge: Page, pagina: int,
                             id_tabela: str = None) -> bool:
    """Navega o paginador DO MODAL para a pagina pedida (1-based), clicando no
    botao numerado. Devolve True se a pagina ativa terminou sendo a pedida."""
    id_tabela = id_tabela or ID_TABELA_RECURSOS
    try:
        container = _container_datatable(rge, id_tabela)
        if not container.count():
            return pagina == 1
        paginador = container.locator(".ui-paginator").first
        if not paginador.count():
            return pagina == 1

        def pagina_ativa() -> int:
            try:
                ativa = paginador.locator(".ui-paginator-page.ui-state-active").first
                return int(ativa.inner_text().strip()) if ativa.count() else 1
            except Exception:
                return 1

        if pagina_ativa() == pagina:
            return True

        botoes = paginador.locator(".ui-paginator-page")
        alvo = None
        for i in range(botoes.count()):
            try:
                if botoes.nth(i).inner_text().strip() == str(pagina):
                    alvo = botoes.nth(i)
                    break
            except Exception:
                continue
        if alvo is None:
            log.warning("Pagina %d nao encontrada no paginador do modal", pagina)
            return False

        alvo.click(timeout=6000)
        aguardar_ajax(rge)
        rge.locator(f'[id="{id_tabela}_data"]').wait_for(timeout=10000)
        ok = pagina_ativa() == pagina
        if not ok:
            log.warning("Cliquei na pagina %d do modal mas a ativa e %d",
                        pagina, pagina_ativa())
        return ok
    except Exception as e:
        log.warning("Falha ao navegar para a pagina %d do modal: %s", pagina, e)
        return False


def _ler_linhas_modal(rge: Page):
    """Le cabecalhos, mapa de colunas e linhas da pagina ATUAL da tabela de
    protocolos do modal. Devolve None quando a tabela nao esta na tela."""
    return rge.evaluate(
        r"""(idTabela) => {
            const tabela = document.getElementById(idTabela)
                        || document.getElementById(idTabela + '_data')?.closest('div.ui-datatable');
            if (!tabela) return null;

            const normalizar = (s) => (s || '')
                .normalize('NFD').replace(/[̀-ͯ]/g, '')
                .replace(/\s+/g, ' ').trim().toLowerCase();

            const cabecalhos = Array.from(tabela.querySelectorAll('thead th'))
                .map(th => normalizar(th.innerText));

            const acharCol = (...alvos) => {
                for (const alvo of alvos) {
                    const i = cabecalhos.findIndex(h => h.includes(alvo));
                    if (i >= 0) return i;
                }
                return -1;
            };

            const col = {
                protocolo: acharCol('protocolo'),
                dataRecurso: acharCol('data recurso', 'data do recurso'),
                dataRetorno: acharCol('complemento', 'data retorno', 'retorno'),
                recursado: acharCol('recursado'),
                qtdItens: acharCol('quantidade de itens', 'qtd de itens', 'itens'),
                inf: acharCol('inf'),
                acatado: acharCol('acatado'),
            };

            const corpo = tabela.querySelector('tbody[id$="_data"]') || tabela.querySelector('tbody');
            const linhas = Array.from(corpo ? corpo.querySelectorAll('tr[data-ri]') : []);

            return {
                cabecalhos,
                colunas: col,
                linhas: linhas.map((tr, i) => {
                    const tds = Array.from(tr.querySelectorAll('td'))
                        .map(td => (td.innerText || '').trim());
                    const pegar = (idx) => (idx >= 0 && idx < tds.length) ? tds[idx] : '';
                    return {
                        indice: parseInt(tr.getAttribute('data-ri'), 10),
                        protocolo: pegar(col.protocolo),
                        dataRecurso: pegar(col.dataRecurso),
                        dataRetorno: pegar(col.dataRetorno),
                        recursado: pegar(col.recursado),
                        qtdItens: pegar(col.qtdItens),
                        inf: pegar(col.inf),
                        acatado: pegar(col.acatado),
                        ehEnvio: !!tr.querySelector('[id*="linkVisualizarRecurso"]'),
                        celulas: tds,
                    };
                }),
            };
        }""",
        ID_TABELA_RECURSOS,
    )

def listar_protocolos(rge: Page):
    """Modal 'Detalhes da Guia'. Cada protocolo aparece em duas linhas: envio
    (link 'linkVisualizarRecurso') e retorno (link 'linkVisualizarRetorno').
    Agrupamos as duas pelo numero do protocolo.

    As colunas sao localizadas pelo CABECALHO, nao por indice fixo: as duas
    abas tem layouts diferentes, e ler a coluna errada produzia valores altos
    demais e repetidos entre guias, sem nenhum erro aparente.

    PAGINACAO: o modal pagina em 10 linhas (envio+retorno = 2 linhas por
    protocolo, entao 15 protocolos = 3 paginas). Primeiro tenta subir o
    paginador para 50; se ainda sobrar pagina, CAMINHA por todas elas lendo
    cada uma. Devolve (lista_de_protocolos, paginas_que_falharam)."""
    expandir_paginador_do_modal(rge, ID_TABELA_RECURSOS)
    total_paginas = paginas_do_modal(rge, ID_TABELA_RECURSOS)

    linhas_todas: List[Dict] = []
    cabecalhos = None
    colunas = None
    paginas_falhas = 0

    for pagina in range(1, total_paginas + 1):
        if total_paginas > 1 and not garantir_pagina_do_modal(rge, pagina):
            paginas_falhas += 1
            continue
        dados = _ler_linhas_modal(rge)
        if not dados:
            if pagina == 1 and total_paginas == 1:
                log.warning("Tabela de recursos nao encontrada no modal")
                return [], 0
            paginas_falhas += 1
            log.warning("Pagina %d do modal sem tabela legivel", pagina)
            continue
        if cabecalhos is None:
            cabecalhos, colunas = dados["cabecalhos"], dados["colunas"]
            faltando = [k for k, v in colunas.items() if v < 0]
            if faltando:
                log.warning("Colunas nao localizadas pelo cabecalho: %s | cabecalhos: %s",
                            faltando, cabecalhos)
        linhas_todas.extend(dados["linhas"])

    if total_paginas > 1:
        log.info("Modal com %d pagina(s): %d linha(s) coletadas, %d pagina(s) com falha",
                 total_paginas, len(linhas_todas), paginas_falhas)

    protocolos: Dict[str, Dict] = {}
    ordem: List[str] = []

    for linha in linhas_todas:
        protocolo = (linha.get("protocolo") or "").strip()
        if not protocolo:
            continue

        if protocolo not in protocolos:
            protocolos[protocolo] = {"protocolo": protocolo, "row_index": None}
            ordem.append(protocolo)
        entry = protocolos[protocolo]

        if linha.get("ehEnvio") and entry["row_index"] is None:
            entry["row_index"] = linha["indice"]

        for chave_origem, chave_destino in (
            ("dataRecurso", "data_recurso"),
            ("dataRetorno", "data_complemento"),
            ("inf", "valor_unit"),
            ("recursado", "valor_total"),
            ("qtdItens", "qtd_itens_protocolo"),
            ("acatado", "valor_acatado"),
        ):
            valor = (linha.get(chave_origem) or "").strip()
            if valor and not entry.get(chave_destino):
                entry[chave_destino] = valor

    return [protocolos[p] for p in ordem], paginas_falhas


def garantir_linha_visivel(rge: Page, row_index: int,
                           id_tabela: str = None) -> bool:
    """Garante que a linha do protocolo (link de envio) esteja renderizada
    antes do clique. O data-ri e um indice ABSOLUTO no modelo de dados, mas o
    link so existe no DOM quando a pagina do modal que contem aquela linha
    esta exibida - e o paginador pode resetar para 10/pagina quando o modal e
    reaberto. Se o link nao estiver na tela, tenta expandir o paginador e, se
    preciso, caminha pelas paginas ate achar."""
    id_tabela = id_tabela or ID_TABELA_RECURSOS
    seletor = f'[id="{id_tabela}:{row_index}:linkVisualizarRecurso"]'
    try:
        if rge.locator(seletor).count():
            return True
        expandir_paginador_do_modal(rge, id_tabela)
        if rge.locator(seletor).count():
            return True
        total = paginas_do_modal(rge, id_tabela)
        for pagina in range(1, total + 1):
            if not garantir_pagina_do_modal(rge, pagina, id_tabela):
                continue
            if rge.locator(seletor).count():
                log.info("Linha do protocolo (row %d) encontrada na pagina %d do modal",
                         row_index, pagina)
                return True
        return False
    except Exception as e:
        log.warning("Falha ao garantir visibilidade da linha %d: %s", row_index, e)
        return False


def abrir_visualizar_protocolo(rge: Page, row_index: int, voltar: bool = True) -> Dict:
    """Abre a tela 'Visualizar protocolo' e, ao final, tenta VOLTAR pelo
    historico do navegador.

    Isso e o que torna o processo rapido: sem voltar, cada protocolo obrigava a
    refazer login->aba->filtros->pesquisa->modal. Voltando, a busca e o modal
    continuam validos e o proximo protocolo e so mais um clique.
    O chamador confere se deu certo (funcao voltou_para_modal) e refaz a busca
    apenas quando o retorno falha."""
    if not garantir_linha_visivel(rge, row_index):
        raise RuntimeError(
            f"Linha do protocolo (row {row_index}) nao esta visivel em "
            f"nenhuma pagina do modal - paginacao do modal falhou"
        )
    rge.locator(f'[id="{ID_TABELA_RECURSOS}:{row_index}:linkVisualizarRecurso"]').click()
    aguardar_ajax(rge)
    aguardar_pagina_pronta(rge)

    data_uso = ""
    for id_data in IDS_DATA_REALIZACAO:
        try:
            alvo = rge.locator(f'[id="{id_data}"]')
            if alvo.count():
                data_uso = alvo.first.inner_text(timeout=5000).strip()
                if data_uso:
                    break
        except Exception:
            continue
    if not data_uso:
        log.warning("Nao consegui ler 'Data da realizacao'")

    itens = []
    try:
        try:
            rge.locator(f'[id="{ID_TABELA_ITENS}"]').wait_for(timeout=8000)
        except Exception:
            rge.wait_for_timeout(1500)   # aba Itens usa outro ID

        brutos = rge.evaluate(
            """(idTabela) => {
                let tbody = document.getElementById(idTabela + '_data');
                if (!tbody) {
                    // aba diferente pode usar outro id: procura pela coluna
                    for (const d of document.querySelectorAll('div.ui-datatable')) {
                        if (/C.d\\.?\\s*Servi/i.test(d.innerText || '')) {
                            tbody = d.querySelector('tbody[id$="_data"]');
                            if (tbody) break;
                        }
                    }
                }
                if (!tbody) return null;

                const linhas = Array.from(tbody.querySelectorAll('tr'));
                const saida = [];
                for (let i = 0; i < linhas.length; i++) {
                    const tr = linhas[i];
                    if (!tr.hasAttribute('data-ri')) continue;
                    const celulas = Array.from(tr.querySelectorAll('td'))
                        .map(td => (td.innerText || '').trim());

                    // A justificativa pode estar numa celula desta linha (Mat/Med)
                    // ou numa linha irma escondida logo abaixo (Itens).
                    let justificativa = '';
                    const prox = linhas[i + 1];
                    if (prox && !prox.hasAttribute('data-ri') &&
                        /justificativa/i.test(prox.innerText || '')) {
                        justificativa = prox.innerText || '';
                    }
                    saida.push({ celulas, justificativa });
                }
                return saida;
            }""",
            ID_TABELA_ITENS,
        )

        if brutos is None:
            caminho = capturar_screenshot_erro(rge, "sem_tabela_itens")
            raise RuntimeError(f"Tabela de itens nao encontrada. Print: {caminho}")

        for bruto in brutos:
            itens.append(_extrair_item(bruto["celulas"], bruto.get("justificativa", "")))
    except Exception as e:
        log.warning("Nao consegui ler a tabela de itens: %s", e)

    # Lido sempre: mesmo com itens, a justificativa deles pode vir vazia (o
    # portal usa "-" como placeholder) e ai o texto util esta na "Observacao"
    # do protocolo.
    protocolo_info = _dados_do_protocolo(rge)

    voltou = False
    if voltar:
        # go_back() nao funciona aqui: o JSF guarda estado no servidor e voltar
        # pelo historico devolve uma pagina expirada. A propria tela tem um
        # botao "Voltar", que o portal trata corretamente.
        for seletor in (
            'button:has-text("Voltar")',
            'input[value="Voltar"]',
            'a:has-text("Voltar")',
        ):
            try:
                alvo = rge.locator(seletor)
                if not alvo.count():
                    continue
                alvo.first.click(timeout=6000)
                aguardar_ajax(rge)
                aguardar_pagina_pronta(rge)
                voltou = busca_ainda_valida(rge)
                if voltou:
                    break
            except Exception:
                continue
        if not voltou:
            log.info("Retorno rapido indisponivel; a busca sera refeita")

    return {
        "data_uso": data_uso,
        "itens": itens,
        "protocolo_info": protocolo_info,
        "voltou": voltou,
    }


def busca_ainda_valida(rge: Page) -> bool:
    """Depois de "Voltar", o portal devolve a TELA DE RESULTADOS - o modal
    fecha, mas a busca continua valida.

    Antes eu verificava se o modal estava aberto; como nunca estava, o robo
    concluia que o retorno falhou e refazia a pesquisa inteira (30-50s) quando
    bastava reabrir o modal da guia (3-5s). Era a causa de "0 retornos rapidos"
    em todas as execucoes."""
    try:
        linhas = rge.locator(f'[id="{ID_TABELA_GUIAS}_data"] tr[data-ri]')
        return linhas.count() > 0 and linhas.first.is_visible()
    except Exception:
        return False


def _dados_do_protocolo(rge: Page) -> Dict:
    """Quando o recurso e feito para a guia inteira (Obj. do recurso = "Guia"),
    a tela nao tem tabela de itens: traz um bloco unico com a justificativa.

    O rotulo aparece de duas formas: sozinho num elemento, com o texto num
    irmao, OU junto do texto no mesmo elemento ("Justificativa: Solicito...").
    Tratamos as duas - so a primeira era coberta antes, e por isso esses
    protocolos vinham sem justificativa nenhuma."""
    try:
        return rge.evaluate(r"""() => {
            const limpar = (s) => (s || '')
                .replace(/\s+/g, ' ').trim();

            const LIXO = /(Voltar|Visualizar Retorno|Copyright|SulAm[ée]rica\s*-\s*\d|Imprimir|Documentos Anexados)/i;
            const cortarLixo = (s) => {
                let t = limpar(s);
                const m = t.match(LIXO);
                if (m && m.index > 0) t = t.slice(0, m.index);
                return t.replace(/[\s.;,-]+$/, '').trim();
            };

            const ROTULO = /^(Justificativa|Observa[çc][ãa]o)\s*:?\s*/i;
            const seletor = 'td, div, span, label, p';

            // Textos que sao ROTULO de outra coisa, nunca justificativa.
            // Sem isso, o elemento seguinte a "Justificativa" na linha do item
            // e o proximo cabecalho ("Valor", "Grau Participacao") e era ele
            // que acabava gravado na planilha.
            const NAO_E_JUSTIFICATIVA = new RegExp(
                '^(valor|grau\\s+participa|documentos?\\s+anexad|itens?\\s+recursad|recurso|' +
                'dados\\s+da\\s+guia|c[oó]d\\.?\\s|qtde|quantidade|aceita\\s+glosa|seq\\.?\\s|' +
                'status|protocolo|data\\s|n[ºo°]\\s|justificativa|observa)', 'i');

            // Uma justificativa real e uma frase: varias palavras e alguma letra.
            const pareceFrase = (s) =>
                s.length >= 12 &&
                (s.match(/\s/g) || []).length >= 2 &&
                /[A-Za-zÀ-ÿ]{3}/.test(s) &&
                !NAO_E_JUSTIFICATIVA.test(s);

            let justificativa = '';
            let menor = Infinity;

            for (const el of document.querySelectorAll(seletor)) {
                const texto = (el.innerText || '').trim();
                if (!ROTULO.test(texto)) continue;

                // a) rotulo e valor no MESMO elemento
                const resto = texto.replace(ROTULO, '').trim();
                if (resto.length > 3) {
                    // o menor elemento que contem o par e o mais especifico:
                    // evita capturar a pagina inteira junto
                    if (texto.length < menor) {
                        const limpo = cortarLixo(resto);
                        if (pareceFrase(limpo)) { justificativa = limpo; menor = texto.length; }
                    }
                    continue;
                }

                // b) rotulo sozinho, valor num elemento irmao
                for (const c of [el.nextElementSibling,
                                 el.parentElement ? el.parentElement.nextElementSibling : null]) {
                    if (!c) continue;
                    const limpo = cortarLixo(c.innerText);
                    if (pareceFrase(limpo) && c.innerText.length < menor) {
                        justificativa = limpo; menor = c.innerText.length;
                        break;
                    }
                }
            }

            const acharValor = (re) => {
                const m = (document.body.innerText || '').match(re);
                return m ? m[1].trim() : '';
            };

            return {
                justificativa,
                valor_total: acharValor(/Valor Total Recursado:?\s*(R\$\s*[\d.,]+)/i),
                objeto: acharValor(new RegExp('Obj\\. do recurso de glosa:?\\s*([^\\n]{1,40})', 'i')),
                num_guia_prestador: acharValor(new RegExp('N[ºo°]\\s*da guia no prestador:?\\s*(\\d+)', 'i')),
            };
        }""")
    except Exception as e:
        log.warning("Nao consegui ler os dados do protocolo: %s", e)
        return {}


def _localizar_tabela_itens(rge: Page):
    """A tabela de itens tem ID diferente em cada aba. Em vez de fixar o do
    Mat/Med, tenta o conhecido e, se nao achar, procura qualquer DataTable que
    tenha a coluna 'Cod. Servico'."""
    alvo = rge.locator(f'[id="{ID_TABELA_ITENS}_data"]')
    if alvo.count():
        return alvo.first

    for texto in ("Cód. Serviço", "Cod. Servico", "Cód. Glosa"):
        candidatos = rge.locator(
            f'div.ui-datatable:has-text("{texto}") tbody[id$="_data"]'
        )
        if candidatos.count():
            tid = candidatos.first.get_attribute("id")
            log.info("Tabela de itens encontrada por conteudo: %s", tid)
            return candidatos.first

    return None


LIXO_RODAPE = re.compile(
    r"(Voltar|Visualizar Retorno|Copyright|SulAm[ée]rica\s*-\s*\d|Imprimir)", re.I
)

# Rotulos que aparecem DEPOIS da justificativa na linha expandida do item.
# Sem cortar aqui, o texto capturado ia ate o fim do bloco e sobrava o "-" do
# Grau Participacao no lugar da justificativa.
ROTULO_SEGUINTE = re.compile(
    r"^(valor|grau\s+participa|c[óo]d\.?\s|qtde|aceita\s+glosa|seq\.?\s*item)\b", re.I
)
ROTULO_JUSTIFICATIVA = re.compile(r"^(justificativa|observa[çc][ãa]o)\s*:?$", re.I)


def _limpar_justificativa(texto: str) -> str:
    """Extrai so o texto da justificativa.

    A linha expandida do item traz varios blocos empilhados:
        Justificativa / <texto> / Valor / R$ X / Grau Participacao / -
    Pegar a ultima linha devolvia o "-" do Grau Participacao. Agora lemos o que
    vem DEPOIS do rotulo e paramos no proximo rotulo conhecido.
    """
    if not texto:
        return ""

    linhas = [l.strip() for l in texto.splitlines() if l.strip()]
    linhas = [l for l in linhas if not LIXO_RODAPE.match(l)]
    if not linhas:
        return ""

    idx = next((i for i, l in enumerate(linhas) if ROTULO_JUSTIFICATIVA.match(l)), None)

    if idx is not None:
        conteudo = []
        for linha in linhas[idx + 1:]:
            if ROTULO_SEGUINTE.match(linha) or ROTULO_JUSTIFICATIVA.match(linha):
                break
            conteudo.append(linha)
        resultado = " ".join(conteudo)
    else:
        # Formato de linha unica: "Justificativa Solicito reanalise..."
        resultado = " ".join(linhas)
        resultado = re.sub(r"^(justificativa|observa[çc][ãa]o)[:\s]*", "", resultado, flags=re.I)

    corte = LIXO_RODAPE.search(resultado)
    if corte and corte.start() > 0:
        resultado = resultado[: corte.start()]

    resultado = re.sub(r"\s{2,}", " ", resultado).strip()

    # "-", "--", "." isolados sao placeholder do portal, nao justificativa
    if len(resultado) <= 2 or not re.search(r"[A-Za-zÀ-ÿ]", resultado):
        return ""

    return resultado


def _extrair_item(textos: List[str], justificativa_externa: str = "") -> Dict:
    """Identifica cada campo pelo CONTEUDO, nao pela posicao. As duas abas tem
    layouts diferentes - a de Itens traz uma coluna vazia a mais entre
    "Cod. Servico" e "Cod. Glosa" - e ler por deslocamento fixo pegava a
    coluna errada em todos os campos.

    Mat/Med: [.., '22', '0000273772 - TIRA...', '1705', 'R$ 3,95', '1', ..]
    Itens  : ['', '15', '60023120 - TAXA...', '', '1714', 'R$ 22,10', '1', '']

    A celula vazia extra na aba Itens e a coluna "Grau Participacao", que nao
    existe no Mat/Med - por isso a leitura por deslocamento fixo pegava a
    coluna errada em todos os campos seguintes.
    """
    def so_digitos(t: str) -> bool:
        return bool(t) and t.replace(".", "").isdigit()

    # 1) a celula do servico e a unica com " - " e sem quebra de linha
    idx_servico = next(
        (i for i, t in enumerate(textos) if " - " in t and "\n" not in t), None
    )
    if idx_servico is None:
        return {
            "seq": "", "codigo": "", "descricao": "", "cod_glosa": "",
            "valor_glosado": "", "qtde": 0,
            "justificativa": _limpar_justificativa(justificativa_externa),
            "_colunas_cruas": textos,
        }

    cod_servico = textos[idx_servico]
    codigo, _, descricao = cod_servico.partition(" - ")

    depois = textos[idx_servico + 1:]

    # 2) o valor glosado e a celula com "R$"
    idx_valor = next((i for i, t in enumerate(depois) if "R$" in t), None)
    valor_glosado = depois[idx_valor] if idx_valor is not None else ""

    # 3) cod. glosa: celula numerica ANTES do valor
    faixa_glosa = depois[:idx_valor] if idx_valor is not None else depois
    cod_glosa = next((t for t in faixa_glosa if so_digitos(t)), "")

    # 4) qtde: primeira celula numerica DEPOIS do valor
    faixa_qtde = depois[idx_valor + 1:] if idx_valor is not None else []
    qtde_txt = next((t for t in faixa_qtde if so_digitos(t)), "")

    # 5) seq: ultima celula numerica ANTES do servico
    seq = ""
    for t in textos[:idx_servico]:
        if so_digitos(t):
            seq = t

    # 6) justificativa: linha irma escondida (Itens) ou celula da propria linha (Mat/Med)
    justificativa = justificativa_externa
    if not justificativa:
        justificativa = next((t for t in textos if "justificativa" in t.lower()), "")

    return {
        "seq": seq,
        "codigo": codigo.strip().lstrip("0") or "0",
        "descricao": descricao.strip(),
        "cod_glosa": cod_glosa,
        "valor_glosado": valor_glosado,
        "qtde": int(qtde_txt) if qtde_txt.isdigit() else 0,
        "justificativa": _limpar_justificativa(justificativa),
        "_colunas_cruas": textos,
    }


def consolidar_itens(itens: List[Dict], protocolo_info: Optional[Dict] = None) -> Dict:
    """Regra confirmada com o cliente: um protocolo vira UMA linha na planilha,
    usando os dados do PRIMEIRO item. Isso e valido porque a justificativa e a
    mesma para todos os itens do protocolo (a soma dos itens forma o valor da
    glosa, que fica no campo 'Valor total').

    Se as justificativas divergirem, a premissa cai por terra e a linha e
    sinalizada para revisao em vez de gravar um dado enganoso."""
    if not itens:
        # Recurso feito para a guia inteira: nao existe tabela de itens, mas a
        # justificativa e o valor estao no bloco do protocolo. Isso e situacao
        # normal do portal, nao erro - por isso nao vai para revisao manual.
        info = protocolo_info or {}
        justificativa = info.get("justificativa", "")
        if justificativa or info.get("valor_total"):
            return {
                "descricao_item": None,
                "codigo_item": None,
                "cod_glosa": None,
                "qtde": None,
                "justificativa": justificativa,
                "revisao_manual": False,
                "erro": f"Recurso no nivel da guia ({info.get('objeto') or 'sem itens'})",
            }
        return {"revisao_manual": True, "erro": "Nenhum item encontrado no protocolo"}

    justificativas = {it["justificativa"] for it in itens if it["justificativa"]}
    base = itens[0]

    # Quando o item nao traz justificativa, cai para a "Observacao" do
    # protocolo - e o mesmo texto que o colaborador copiava a mao.
    justificativa = base["justificativa"] or (protocolo_info or {}).get("justificativa", "")

    resultado = {
        "descricao_item": base["descricao"],
        "codigo_item": base["codigo"],
        "cod_glosa": base["cod_glosa"],
        "qtde": base["qtde"],          # do 1o item, nao a soma (confirmado)
        "justificativa": justificativa,
        "revisao_manual": False,
    }
    if len(justificativas) > 1:
        resultado["revisao_manual"] = True
        lista = " | ".join(sorted(justificativas))
        resultado["erro"] = (
            f"Protocolo com {len(justificativas)} justificativas diferentes "
            f"entre os {len(itens)} itens: {lista}"
        )
    return resultado


# --------------------------------------------------------------- endpoint ----

@app.post("/processar-lote", response_model=List[ItemExtraido])
def processar_lote(req: ProcessarLoteRequest):
    """Percorre as duas abas (Mat/Med e Itens) para o lote informado.

    DESEMPENHO: abrir um protocolo e um submit JSF que tira o robo da tela de
    resultados. Refazer a busca a cada protocolo custava ~18 pesquisas por
    lote. Agora tentamos voltar pelo historico do navegador, o que preserva a
    busca e o modal - a busca so e refeita quando o retorno falha."""
    resultados: List[ItemExtraido] = []
    metricas = {"buscas": 0, "retornos_rapidos": 0, "retornos_lentos": 0}

    espera_fila = time.time()
    _semaforo.acquire()
    fila = time.time() - espera_fila
    if fila > 1:
        log.info("Lote %s aguardou %.0fs na fila", req.lote, fila)

    try:
        return _processar(req, resultados, metricas)
    finally:
        _semaforo.release()


def _processar(req: ProcessarLoteRequest, resultados: List[ItemExtraido],
               metricas: Dict) -> List[ItemExtraido]:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        contexto = browser.new_context()

        # Imagens, fontes e rastreadores nao servem para extrair dados e
        # respondem por boa parte do tempo de carga de cada tela.
        if BLOQUEAR_MIDIA:
            contexto.route(
                "**/*",
                lambda rota: rota.abort()
                if (rota.request.resource_type in ("image", "media", "font")
                    or "datadoghq" in rota.request.url
                    or "google-analytics" in rota.request.url)
                else rota.continue_(),
            )

        page = contexto.new_page()
        inicio_exec = time.time()
        try:
            try:
                fazer_login(page)
                rge = abrir_rge(page)
            except Exception as e:
                raise HTTPException(status_code=502, detail=str(e))

            vistos = set()   # protocolo ja capturado: evita repetir entre tipos

            for aba, tipo_glosa in PASSAGENS:
                rotulo_passagem = f"{aba}" + (f"/{tipo_glosa}" if tipo_glosa else "")
                try:
                    ir_para_aba(rge, aba)
                except Exception as e:
                    log.exception("Falha ao entrar na aba %s", aba)
                    resultados.append(ItemExtraido(
                        lote=req.lote,
                        aba=aba, erro=f"Nao consegui abrir a aba {aba}: {e}",
                        revisao_manual=True,
                    ))
                    continue

                def buscar(_aba=aba, _tipo=tipo_glosa):
                    metricas["buscas"] += 1
                    pesquisar_lote(rge, req.lote, req.data_inicio_pagto,
                                   req.data_fim_pagto, _aba, _tipo)

                try:
                    buscar()
                    guias = listar_guias(rge)
                except SemResultados as e:
                    # Normal: nem todo lote tem recurso em toda combinacao
                    log.info("%s, lote %s: %s", rotulo_passagem, req.lote, e)
                    continue
                except Exception as e:
                    # Falha de verdade precisa aparecer: foi tratar as duas
                    # coisas como iguais que escondeu o Mat/Med inteiro parando.
                    log.warning("%s, lote %s: falha na busca (%s)", rotulo_passagem, req.lote, e)
                    resultados.append(ItemExtraido(
                        lote=req.lote,
                        aba=aba, revisao_manual=True,
                        erro=f"Busca falhou na aba {aba}: {e}",
                    ))
                    continue

                if not guias:
                    log.info("%s, lote %s: nenhuma guia", rotulo_passagem, req.lote)
                    continue

                # Mesmo com o paginador em 50, sobrou mais de uma pagina
                # (>50 guias no lote): as paginas seguintes nao serao lidas.
                # Precisa virar linha na planilha - so log seria invisivel.
                n_paginas = paginas_de_resultado(rge)
                if n_paginas > 1:
                    resultados.append(ItemExtraido(
                        lote=req.lote,
                        aba=aba,
                        protocolo=f"PAGINACAO_{aba.upper()}",
                        erro=(f"Resultado com {n_paginas} paginas na aba {aba}; "
                              f"somente a primeira ({len(guias)} guias) foi lida - "
                              "protocolos das demais paginas NAO capturados"),
                        revisao_manual=True,
                    ))

                log.info("%s, lote %s: %d guia(s)", rotulo_passagem, req.lote, len(guias))

                for guia in guias:
                    rotulo = guia.get("guia", "")

                    # Abrir a guia e listar seus protocolos e o unico passo do
                    # loop que, se falhar, perde TODOS os protocolos da guia
                    # de uma vez - diferente da falha por protocolo (mais
                    # abaixo), que ja tinha tratamento individual. Uma falha
                    # aqui costuma ser transitoria (elemento ainda carregando
                    # logo apos trocar de guia/tipo de glosa), entao vale
                    # tentar de novo antes de desistir - mesmo padrao usado
                    # em pesquisar_lote e abrir_rge.
                    protocolos = None
                    paginas_falhas = 0
                    erro_abertura = None
                    for tentativa in (1, 2):
                        try:
                            abrir_detalhes_guia(rge, guia["row_index"], rotulo)
                            protocolos, paginas_falhas = listar_protocolos(rge)
                            if paginas_falhas > 0:
                                # Alguma pagina do modal nao pode ser lida
                                # mesmo apos expandir/navegar o paginador.
                                # Nao ha como saber quais protocolos ficaram
                                # de fora - registra em vez de fingir que
                                # leu tudo.
                                resultados.append(ItemExtraido(
                                    lote=req.lote,
                                    aba=aba, guia=rotulo,
                                    protocolo=f"PAGINACAO_MODAL_{rotulo}",
                                    erro=(f"{paginas_falhas} pagina(s) do modal de protocolos "
                                          f"da guia {rotulo} nao puderam ser lidas; "
                                          f"{len(protocolos)} protocolos coletados nas demais"),
                                    revisao_manual=True,
                                ))
                            break
                        except Exception as e:
                            erro_abertura = e
                            log.warning(
                                "Falha ao abrir guia %s (aba %s), tentativa %d/2: %s",
                                rotulo, aba, tentativa, e,
                            )
                            if tentativa == 1:
                                pausa_humana(rge)

                    if protocolos is None:
                        # Sem lista de protocolos nao ha como saber quais
                        # numeros ficaram de fora - mas o registro precisa
                        # sobreviver ate a planilha. Sem campo "protocolo" o
                        # /preencher-planilha descarta a linha em silencio
                        # (contador "ignoradas"), entao o guia+lote+erro fica
                        # so no log do servidor. Prefixar com o rotulo da
                        # guia garante uma linha rastreavel na planilha,
                        # mesmo sem o(s) numero(s) de protocolo.
                        capturar_screenshot_erro(rge, f"guia_{rotulo}_{aba}")
                        resultados.append(ItemExtraido(
                            lote=req.lote,
                            aba=aba, guia=rotulo,
                            protocolo=f"GUIA_{rotulo}_SEM_ABRIR",
                            erro=f"Nao consegui abrir a guia apos 2 tentativas: {erro_abertura}",
                            revisao_manual=True,
                        ))
                        continue

                    try:
                        log.info("  guia %s: %d protocolo(s)", rotulo, len(protocolos))
                        if protocolos and all(
                            str(p.get("protocolo") or "").strip() in vistos for p in protocolos
                        ):
                            log.warning(
                                "  guia %s (aba %s): TODOS os %d protocolos lidos ja estao em "
                                "'vistos' - suspeita de tabela com conteudo da abertura anterior",
                                rotulo, aba, len(protocolos),
                            )

                        # o modal ja esta aberto para o 1o protocolo
                        primeiro = True
                        busca_valida = False
                        for prot in protocolos:
                            # Um protocolo pode aparecer na busca administrativa
                            # e na tecnica; grava so na primeira vez.
                            # O registro em "vistos" fica para DEPOIS de dar
                            # certo: marcar antes fazia um protocolo que falhou
                            # na passagem administrativa ser pulado na tecnica
                            # e em Itens - jogando fora a segunda chance que as
                            # 3 passagens dao de graca.
                            chave = str(prot.get("protocolo") or "").strip()
                            if chave and chave in vistos:
                                continue

                            if prot.get("row_index") is None:
                                resultados.append(ItemExtraido(
                                    lote=req.lote,
                                    aba=aba, guia=rotulo, protocolo=prot["protocolo"],
                                    data_recurso=prot.get("data_recurso"),
                                    data_complemento=prot.get("data_complemento"),
                                    valor_unit=prot.get("valor_unit"),
                                    valor_total=prot.get("valor_total"),
                                    revisao_manual=True,
                                    erro="Protocolo sem linha de envio (so retorno)",
                                ))
                                continue

                            try:
                                # Caminho rapido: a busca sobreviveu ao "Voltar",
                                # entao basta reabrir o modal da guia (3-5s) em
                                # vez de refazer a pesquisa toda (30-50s).
                                if primeiro:
                                    primeiro = False          # modal ja aberto
                                elif busca_valida:
                                    abrir_detalhes_guia(rge, guia["row_index"], rotulo)
                                    metricas["retornos_rapidos"] += 1
                                else:
                                    buscar()
                                    abrir_detalhes_guia(rge, guia["row_index"], rotulo)
                                    metricas["retornos_lentos"] += 1

                                detalhe = abrir_visualizar_protocolo(rge, prot["row_index"])
                                busca_valida = detalhe.get("voltou", False)

                                consolidado = consolidar_itens(
                                    detalhe["itens"], detalhe.get("protocolo_info")
                                )
                                resultados.append(ItemExtraido(
                                    lote=req.lote,
                                    aba=aba,
                                    guia=rotulo,
                                    protocolo=prot["protocolo"],
                                    data_recurso=prot.get("data_recurso"),
                                    data_complemento=prot.get("data_complemento"),
                                    valor_unit=prot.get("valor_unit"),
                                    valor_total=(prot.get("valor_total")
                                                 or (detalhe.get("protocolo_info") or {}).get("valor_total")),
                                    valor_acatado=prot.get("valor_acatado"),
                                    qtd_itens_protocolo=prot.get("qtd_itens_protocolo"),
                                    data_uso=detalhe["data_uso"],
                                    **consolidado,
                                ))
                                # Só agora o protocolo conta como capturado.
                                if chave:
                                    vistos.add(chave)
                            except Exception as e:
                                # Nao entra em "vistos": se a falha foi
                                # transitoria, a proxima passagem (tecnica ou
                                # Itens) tenta o mesmo protocolo de novo. Se
                                # falhar em todas, sobra a linha REVISAR - a
                                # ultima sobrescreve a anterior na planilha,
                                # entao nao gera duplicata.
                                log.exception("Falha no protocolo %s", prot.get("protocolo"))
                                capturar_screenshot_erro(rge, f"prot_{prot.get('protocolo')}")
                                resultados.append(ItemExtraido(
                                    lote=req.lote,
                                    aba=aba, guia=rotulo,
                                    protocolo=prot.get("protocolo", ""),
                                    erro=str(e), revisao_manual=True,
                                ))
                                busca_valida = False
                    except Exception as e:
                        # Falha inesperada em algum ponto do loop de
                        # protocolos que nao foi pega pelo try/except por
                        # protocolo acima (ex.: erro na propria iteracao).
                        # Mesmo raciocinio do bloco de abertura da guia: sem
                        # o campo "protocolo", a linha e descartada em
                        # silencio pelo /preencher-planilha. Como aqui ja
                        # temos a lista completa de "protocolos" da guia,
                        # gravamos um REVISAR para cada um que ainda nao foi
                        # registrado em "vistos" - os ja processados antes
                        # da falha nao sao duplicados.
                        log.exception("Falha na guia %s (aba %s)", rotulo, aba)
                        pendentes_da_guia = [
                            p for p in (protocolos or [])
                            if str(p.get("protocolo") or "").strip() not in vistos
                        ]
                        if pendentes_da_guia:
                            for p in pendentes_da_guia:
                                chave = str(p.get("protocolo") or "").strip()
                                if chave:
                                    vistos.add(chave)
                                resultados.append(ItemExtraido(
                                    lote=req.lote,
                                    aba=aba, guia=rotulo,
                                    protocolo=chave or f"GUIA_{rotulo}_FALHA_NO_LOOP",
                                    erro=f"Falha na guia (protocolo nao processado): {e}",
                                    revisao_manual=True,
                                ))
                        else:
                            resultados.append(ItemExtraido(
                                lote=req.lote,
                                aba=aba, guia=rotulo,
                                protocolo=f"GUIA_{rotulo}_FALHA_NO_LOOP",
                                erro=str(e), revisao_manual=True,
                            ))
        finally:
            browser.close()

    duracao = time.time() - inicio_exec
    log.info(
        "Lote %s: %d registro(s) em %.0fs | %d busca(s), %d retorno(s) rapido(s), %d lento(s)",
        req.lote, len(resultados), duracao,
        metricas["buscas"], metricas["retornos_rapidos"], metricas["retornos_lentos"],
    )
    return resultados


@app.get("/health")
def health():
    return {"status": "ok"}


# ------------------------------------------------- preenchimento da planilha ----
#
# Recriar a planilha a partir de JSON (como o no "JSON -> Excel" do n8n faz)
# destroi formatacao de data, formato monetario, larguras de coluna, estilos e
# abas extras. Aqui abrimos o arquivo ORIGINAL e escrevemos apenas nas celulas
# necessarias, preservando todo o resto.

import json
from io import BytesIO
from copy import copy
from datetime import datetime

import openpyxl
from fastapi import File, Form, UploadFile
from fastapi.responses import StreamingResponse

LINHA_CABECALHO = 2
PRIMEIRA_LINHA_DADOS = 3
COL_LOTE_PLANILHA = "PROTOCOLO DE ENTREGA"
COL_PROTOCOLO = "PROTOCOLO RECURSO"
COL_STATUS = "Status Robô"


def _moeda(v):
    if v in (None, ""):
        return None
    limpo = re.sub(r"[^\d,.-]", "", str(v)).replace(".", "").replace(",", ".")
    try:
        return float(limpo)
    except ValueError:
        return None


def _data(v):
    if not v:
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", str(v).strip())
    return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else None


def _num(v):
    if v in (None, ""):
        return None
    try:
        return int(str(v).strip())
    except ValueError:
        return v


def _mapa_colunas(ws) -> Dict[str, int]:
    return {
        ws.cell(row=LINHA_CABECALHO, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=LINHA_CABECALHO, column=c).value
    }


def _garantir_coluna_status(ws, colunas: Dict[str, int]) -> int:
    if COL_STATUS in colunas:
        return colunas[COL_STATUS]
    nova = ws.max_column + 1
    celula = ws.cell(row=LINHA_CABECALHO, column=nova)
    celula.value = COL_STATUS
    modelo = ws.cell(row=LINHA_CABECALHO, column=1)
    celula.font = copy(modelo.font)
    celula.fill = copy(modelo.fill)
    celula.border = copy(modelo.border)
    celula.alignment = copy(modelo.alignment)
    colunas[COL_STATUS] = nova
    return nova


COLUNAS_DATA = {"DATA DO RECURSO", "Data uso", "Data Recurso 1", "Data Retorno 1",
                "Data Recurso 2", "Data Retorno 2"}
COLUNAS_MOEDA = {"Valor Unit", "Valor total", "Valor Recurso", "Vl Recuperado",
                 "Valor Pendente"}


def _formatos_por_coluna(ws, colunas: Dict[str, int]) -> Dict[str, str]:
    """Descobre o formato que a planilha ja usa em cada coluna. Celulas vazias
    nao tem formato, entao escrever nelas sem definir number_format faria a
    data aparecer como numero de serie (o 45742 que voce viu)."""
    formatos = {}
    for nome, col in colunas.items():
        # Datas sempre em dd/mm/aaaa: a planilha original mistura dd/mm/yy,
        # mm-dd-yy e texto na mesma coluna, e herdar isso deixaria o resultado
        # dependente do idioma do Excel de quem abrir.
        if nome in COLUNAS_DATA:
            formatos[nome] = "dd/mm/yyyy"
            continue
        encontrado = None
        for r in range(PRIMEIRA_LINHA_DADOS, min(ws.max_row, PRIMEIRA_LINHA_DADOS + 400) + 1):
            celula = ws.cell(row=r, column=col)
            if celula.value is not None and celula.number_format != "General":
                encontrado = celula.number_format
                break
        if not encontrado:
            if nome in COLUNAS_DATA:
                encontrado = "dd/mm/yy"
            elif nome in COLUNAS_MOEDA:
                encontrado = "#,##0.00"
        if encontrado:
            formatos[nome] = encontrado
    return formatos


def _copiar_estilo_da_linha(ws, origem: int, destino: int, ultima_coluna: int):
    for c in range(1, ultima_coluna + 1):
        de = ws.cell(row=origem, column=c)
        para = ws.cell(row=destino, column=c)
        para.font = copy(de.font)
        para.fill = copy(de.fill)
        para.border = copy(de.border)
        para.alignment = copy(de.alignment)
        para.number_format = de.number_format


@app.post("/preencher-planilha")
async def preencher_planilha(arquivo: UploadFile = File(...), linhas: str = Form(...)):
    """Recebe a planilha original + os protocolos extraidos e devolve o mesmo
    arquivo preenchido, com formatacao intacta."""
    resultados = json.loads(linhas)
    if isinstance(resultados, dict):
        resultados = [resultados]

    wb = openpyxl.load_workbook(BytesIO(await arquivo.read()))
    ws = wb["Modificado"] if "Modificado" in wb.sheetnames else wb[wb.sheetnames[0]]

    colunas = _mapa_colunas(ws)
    formatos = _formatos_por_coluna(ws, colunas)
    col_status = _garantir_coluna_status(ws, colunas)
    ultima_coluna = ws.max_column

    # Onde termina o grupo de cada lote e em que linha esta cada protocolo
    fim_do_lote: Dict[str, int] = {}
    linha_do_protocolo: Dict[str, int] = {}
    lote_atual = None
    for r in range(PRIMEIRA_LINHA_DADOS, ws.max_row + 1):
        lote = ws.cell(row=r, column=colunas[COL_LOTE_PLANILHA]).value
        if lote not in (None, ""):
            lote_atual = str(lote).strip()
        if lote_atual:
            fim_do_lote[lote_atual] = r
        prot = ws.cell(row=r, column=colunas[COL_PROTOCOLO]).value
        if prot not in (None, ""):
            linha_do_protocolo[str(prot).strip()] = r

    atualizadas = criadas = ignoradas = 0

    for res in resultados:
        protocolo = str(res.get("protocolo") or "").strip()
        if not protocolo:
            ignoradas += 1
            continue

        # Protocolo sem NENHUM dado util nao vira linha: so gera ruido na
        # planilha, marcado como OK sem conteudo algum.
        tem_conteudo = any(
            res.get(campo) for campo in (
                "data_recurso", "data_complemento", "valor_unit", "valor_total",
                "data_uso", "descricao_item", "codigo_item", "cod_glosa",
                "qtde", "justificativa",
            )
        )
        if not tem_conteudo and not res.get("erro"):
            ignoradas += 1
            log.info("Protocolo %s ignorado: sem dados", protocolo)
            continue

        destino = linha_do_protocolo.get(protocolo)
        if destino is None:
            # Protocolo novo: entra logo apos a ultima linha do lote, para
            # ficar junto dos irmaos em vez de solto no fim da planilha.
            lote_res = str(res.get("lote") or "").strip()
            ancora = fim_do_lote.get(lote_res, ws.max_row)
            destino = ancora + 1
            ws.insert_rows(destino)
            _copiar_estilo_da_linha(ws, ancora, destino, ultima_coluna)
            # Os indices abaixo do ponto de insercao andam uma linha
            for k, v in list(linha_do_protocolo.items()):
                if v >= destino:
                    linha_do_protocolo[k] = v + 1
            for k, v in list(fim_do_lote.items()):
                if v >= destino:
                    fim_do_lote[k] = v + 1
            if lote_res in fim_do_lote:
                fim_do_lote[lote_res] = destino
            linha_do_protocolo[protocolo] = destino
            criadas += 1
        else:
            atualizadas += 1

        valor_recurso = _moeda(res.get("valor_total"))
        valor_recuperado = _moeda(res.get("valor_acatado"))
        pendente = (
            round(valor_recurso - valor_recuperado, 2)
            if valor_recurso is not None and valor_recuperado is not None
            else None
        )

        valores = {
            "Guia": res.get("guia"),
            COL_PROTOCOLO: _num(protocolo),
            "DATA DO RECURSO": _data(res.get("data_recurso")),
            "Data uso": _data(res.get("data_uso")),
            "Descrição Item": res.get("descricao_item"),
            "Codigo Item": _num(res.get("codigo_item")),
            "Valor Unit": _moeda(res.get("valor_unit")),
            "Valor total": valor_recurso,
            "Qtde": res.get("qtde"),
            "Cod glosa": _num(res.get("cod_glosa")),
            "Justificativa recurso": res.get("justificativa"),
            "Valor Recurso": valor_recurso,
            "Vl Recuperado": valor_recuperado,
            "Valor Pendente": pendente,
            "Data Recurso 1": _data(res.get("data_recurso")),
            "Data Retorno 1": _data(res.get("data_complemento")),
        }

        for nome, valor in valores.items():
            col = colunas.get(nome)
            if col and valor is not None:
                celula = ws.cell(row=destino, column=col)
                celula.value = valor
                if nome in formatos:
                    celula.number_format = formatos[nome]

        motivo = res.get("erro")
        ws.cell(row=destino, column=col_status).value = (
            f"REVISAR: {motivo}" if res.get("revisao_manual") and motivo
            else ("REVISAR MANUALMENTE" if res.get("revisao_manual") else "OK")
        )

    # Totais da linha 1 acompanhando o novo tamanho da planilha
    for letra in ("E", "G", "H", "J"):
        celula = ws[f"{letra}1"]
        if celula.value is not None:
            celula.value = f"=SUBTOTAL(9,{letra}{PRIMEIRA_LINHA_DADOS}:{letra}{ws.max_row})"

    log.info("Planilha preenchida: %d atualizada(s), %d criada(s), %d ignorada(s)",
             atualizadas, criadas, ignoradas)

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return StreamingResponse(
        saida,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="planilha_preenchida.xlsx"'},
    )
